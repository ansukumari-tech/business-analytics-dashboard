"""
generate_excel.py
Generates a fully formatted Excel workbook:
  - Sheet 1: Sales Data (raw fact table)
  - Sheet 2: Monthly KPI Dashboard
  - Sheet 3: Category Analysis
  - Sheet 4: Regional Performance
  - Sheet 5: Rep Leaderboard
  - Sheet 6: Product Performance
  - Sheet 7: RFM Customer Table
Run after etl_pipeline.py
"""
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                              numbers as num_styles)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.chart.series import SeriesLabel
from pathlib import Path

RAW       = Path(__file__).parent.parent / "data" / "raw"
PROCESSED = Path(__file__).parent.parent / "data" / "processed"
EXCEL_OUT = Path(__file__).parent.parent / "excel"
EXCEL_OUT.mkdir(exist_ok=True)

# ── Load data ────────────────────────────────────────────────
sales    = pd.read_csv(RAW / "sales_raw.csv", parse_dates=["order_date"])
monthly  = pd.read_csv(PROCESSED / "monthly_summary.csv")
category = pd.read_csv(PROCESSED / "category_summary.csv")
region   = pd.read_csv(PROCESSED / "region_summary.csv")
products = pd.read_csv(PROCESSED / "product_performance.csv")
reps     = pd.read_csv(PROCESSED / "rep_performance.csv")

delivered = sales[sales["status"] == "Delivered"]

# ── Style helpers ─────────────────────────────────────────────
def hdr(cell, text, fg="FFFFFF", bg="3A3A8C", bold=True, sz=11):
    cell.value = text
    cell.font  = Font(bold=bold, color=fg, size=sz, name="Calibri")
    cell.fill  = PatternFill("solid", start_color=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

def val(cell, v, num_fmt=None, bold=False, color="000000"):
    cell.value = v
    cell.font  = Font(bold=bold, color=color, name="Calibri", size=10)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    if num_fmt:
        cell.number_format = num_fmt

def thin_border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

def set_col_widths(ws, widths):
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

def alt_row(ws, row, cols, bg="F0F0FA"):
    for c in cols:
        ws[f"{c}{row}"].fill = PatternFill("solid", start_color=bg)

wb = Workbook()

# ════════════════════════════════════════════════════════════
# SHEET 1: Sales Data
# ════════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "Sales Data"
ws1.freeze_panes = "A2"

cols = ["order_id","order_date","year","month_name","quarter","status",
        "customer_name","segment","city","region_name","rep_name",
        "product_name","category","quantity","unit_price","discount","revenue","cost","profit","margin_pct"]
headers = ["Order ID","Order Date","Year","Month","Quarter","Status",
           "Customer","Segment","City","Region","Sales Rep",
           "Product","Category","Qty","Unit Price","Discount","Revenue","Cost","Profit","Margin %"]

for j, (c, h) in enumerate(zip(cols, headers), 1):
    hdr(ws1.cell(1, j), h)

NUM_FMT_INR = '#,##0'
NUM_FMT_PCT = '0.00%'

for i, row in enumerate(sales[cols].itertuples(index=False), 2):
    bg = "FAFAFA" if i % 2 == 0 else "FFFFFF"
    for j, v in enumerate(row, 1):
        c = ws1.cell(i, j)
        c.value = v
        c.font = Font(size=9, name="Calibri")
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.fill = PatternFill("solid", start_color=bg)
        c.border = thin_border()
        if j in (15, 17, 18, 19):
            c.number_format = '#,##0'
        if j == 20:
            c.number_format = '0.00'

ws1.row_dimensions[1].height = 28
set_col_widths(ws1, {
    "A":9,"B":12,"C":7,"D":8,"E":8,"F":11,"G":22,"H":15,"I":12,
    "J":10,"K":15,"L":22,"M":12,"N":6,"O":11,"P":9,"Q":12,"R":12,"S":12,"T":9
})

ws1.auto_filter.ref = f"A1:{get_column_letter(len(cols))}1"
print("  ✅ Sheet 1: Sales Data")

# ════════════════════════════════════════════════════════════
# SHEET 2: Monthly KPI Dashboard
# ════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("Monthly KPI")
ws2.freeze_panes = "A3"

# KPI Summary row
kpi_labels = ["Total Revenue","Total Profit","Avg Margin %","Total Orders"]
kpi_vals   = [
    delivered["revenue"].sum(),
    delivered["profit"].sum(),
    delivered["profit"].sum() / delivered["revenue"].sum() * 100,
    delivered["order_id"].nunique(),
]
kpi_fmts   = ["#,##0","#,##0","0.00","#,##0"]
kpi_colors = ["6C63FF","43E97B","F7971E","FF6584"]

for j, (lbl, v, fmt, col) in enumerate(zip(kpi_labels, kpi_vals, kpi_fmts, kpi_colors), 2):
    ws2.cell(1, j).value = lbl
    ws2.cell(1, j).font  = Font(bold=True, size=10, color="FFFFFF", name="Calibri")
    ws2.cell(1, j).fill  = PatternFill("solid", start_color=col)
    ws2.cell(1, j).alignment = Alignment(horizontal="center")
    ws2.cell(2, j).value = v
    ws2.cell(2, j).font  = Font(bold=True, size=13, color=col, name="Calibri")
    ws2.cell(2, j).number_format = fmt
    ws2.cell(2, j).alignment = Alignment(horizontal="center")

# Monthly table
mheaders = ["Year","Month","Month","Orders","Revenue (₹)","Cost (₹)","Profit (₹)","Margin %"]
for j, h in enumerate(mheaders, 1):
    hdr(ws2.cell(4, j), h)

monthly_sorted = monthly.sort_values(["year","month"])
for i, row in enumerate(monthly_sorted.itertuples(index=False), 5):
    bg = "F0F0FA" if i % 2 == 0 else "FFFFFF"
    data = [row.year, row.month, row.month_name, row.orders,
            row.revenue, row.cost, row.profit, row.margin_pct]
    fmts = [None,None,None,"#,##0","#,##0","#,##0","#,##0","0.00"]
    for j, (v, fmt) in enumerate(zip(data, fmts), 1):
        c = ws2.cell(i, j)
        c.value = v
        c.font = Font(size=10, name="Calibri")
        c.fill = PatternFill("solid", start_color=bg)
        c.alignment = Alignment(horizontal="center")
        c.border = thin_border()
        if fmt: c.number_format = fmt

# Line chart
last_row = 4 + len(monthly_sorted)
chart = LineChart()
chart.title = "Monthly Revenue vs Profit"
chart.style = 10
chart.height = 12
chart.width  = 22
rev_ref  = Reference(ws2, min_col=5, min_row=4, max_row=last_row)
prof_ref = Reference(ws2, min_col=7, min_row=4, max_row=last_row)
labels   = Reference(ws2, min_col=3, min_row=5, max_row=last_row)
chart.add_data(rev_ref, titles_from_data=True)
chart.add_data(prof_ref, titles_from_data=True)
chart.set_categories(labels)
ws2.add_chart(chart, "J4")

set_col_widths(ws2, {"A":7,"B":7,"C":8,"D":8,"E":14,"F":14,"G":14,"H":10})
ws2.row_dimensions[1].height = 22
ws2.row_dimensions[4].height = 28
print("  ✅ Sheet 2: Monthly KPI")

# ════════════════════════════════════════════════════════════
# SHEET 3: Category Analysis
# ════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("Category Analysis")
cat_headers = ["Category","Revenue (₹)","Profit (₹)","Units Sold","Margin %"]
for j, h in enumerate(cat_headers, 1):
    hdr(ws3.cell(1, j), h)

cat_sorted = category.sort_values("revenue", ascending=False)
for i, row in enumerate(cat_sorted.itertuples(index=False), 2):
    bg = "F0F0FA" if i % 2 == 0 else "FFFFFF"
    data  = [row.category, row.revenue, row.profit, row.units, row.margin_pct]
    fmts  = [None,"#,##0","#,##0","#,##0","0.00"]
    for j, (v, fmt) in enumerate(zip(data, fmts), 1):
        c = ws3.cell(i, j)
        c.value = v
        c.font = Font(size=10, name="Calibri")
        c.fill = PatternFill("solid", start_color=bg)
        c.alignment = Alignment(horizontal="center")
        c.border = thin_border()
        if fmt: c.number_format = fmt

# Totals row
t_row = 2 + len(cat_sorted)
ws3.cell(t_row, 1).value = "TOTAL"
ws3.cell(t_row, 1).font  = Font(bold=True, name="Calibri")
ws3.cell(t_row, 2).value = f"=SUM(B2:B{t_row-1})"
ws3.cell(t_row, 2).number_format = "#,##0"
ws3.cell(t_row, 2).font  = Font(bold=True, color="6C63FF", name="Calibri")
ws3.cell(t_row, 3).value = f"=SUM(C2:C{t_row-1})"
ws3.cell(t_row, 3).number_format = "#,##0"
ws3.cell(t_row, 3).font  = Font(bold=True, color="43E97B", name="Calibri")

# Bar chart
bc = BarChart()
bc.type  = "col"
bc.title = "Revenue by Category"
bc.style = 10
bc.height = 12
bc.width  = 18
rev_ref = Reference(ws3, min_col=2, min_row=1, max_row=t_row-1)
cats    = Reference(ws3, min_col=1, min_row=2, max_row=t_row-1)
bc.add_data(rev_ref, titles_from_data=True)
bc.set_categories(cats)
ws3.add_chart(bc, "G2")

set_col_widths(ws3, {"A":18,"B":14,"C":14,"D":12,"E":10})
print("  ✅ Sheet 3: Category Analysis")

# ════════════════════════════════════════════════════════════
# SHEET 4: Regional Performance
# ════════════════════════════════════════════════════════════
ws4 = wb.create_sheet("Regional Performance")
reg_headers = ["Region","Revenue (₹)","Profit (₹)","Orders","Revenue Rank"]
for j, h in enumerate(reg_headers, 1):
    hdr(ws4.cell(1, j), h)

reg_sorted = region.sort_values("revenue", ascending=False).reset_index(drop=True)
for i, row in enumerate(reg_sorted.itertuples(index=False), 2):
    bg = "F0F0FA" if i % 2 == 0 else "FFFFFF"
    data = [row.region_name, row.revenue, row.profit, row.orders, i-1]
    fmts = [None,"#,##0","#,##0","#,##0",None]
    for j, (v, fmt) in enumerate(zip(data, fmts), 1):
        c = ws4.cell(i, j)
        c.value = v
        c.font = Font(size=10, name="Calibri")
        c.fill = PatternFill("solid", start_color=bg)
        c.alignment = Alignment(horizontal="center")
        c.border = thin_border()
        if fmt: c.number_format = fmt

# Pie chart
pc = PieChart()
pc.title = "Revenue by Region"
pc.style = 10
pc.height = 12
pc.width  = 16
rev_ref = Reference(ws4, min_col=2, min_row=1, max_row=1+len(reg_sorted))
cats    = Reference(ws4, min_col=1, min_row=2, max_row=1+len(reg_sorted))
pc.add_data(rev_ref, titles_from_data=True)
pc.set_categories(cats)
ws4.add_chart(pc, "G2")

set_col_widths(ws4, {"A":14,"B":14,"C":14,"D":10,"E":12})
print("  ✅ Sheet 4: Regional Performance")

# ════════════════════════════════════════════════════════════
# SHEET 5: Rep Leaderboard
# ════════════════════════════════════════════════════════════
ws5 = wb.create_sheet("Rep Leaderboard")
rep_headers = ["Rank","Rep Name","Region","Orders","Revenue (₹)"]
for j, h in enumerate(rep_headers, 1):
    hdr(ws5.cell(1, j), h)

reps_sorted = reps.sort_values("revenue", ascending=False).reset_index(drop=True)
medal_colors = ["FFD700","C0C0C0","CD7F32"]
for i, row in enumerate(reps_sorted.itertuples(index=False), 2):
    bg = medal_colors[i-2] if i <= 4 else ("F0F0FA" if i % 2 == 0 else "FFFFFF")
    rank_color = "8B0000" if i - 1 == 1 else "000000"
    data = [i-1, row.rep_name, row.region_name, row.orders, row.revenue]
    fmts = [None,None,None,"#,##0","#,##0"]
    for j, (v, fmt) in enumerate(zip(data, fmts), 1):
        c = ws5.cell(i, j)
        c.value = v
        c.font = Font(size=10, bold=(j==1), color=rank_color if j==1 else "000000", name="Calibri")
        c.fill = PatternFill("solid", start_color=bg)
        c.alignment = Alignment(horizontal="center")
        c.border = thin_border()
        if fmt: c.number_format = fmt

set_col_widths(ws5, {"A":7,"B":18,"C":12,"D":10,"E":14})
print("  ✅ Sheet 5: Rep Leaderboard")

# ════════════════════════════════════════════════════════════
# SHEET 6: Top Products
# ════════════════════════════════════════════════════════════
ws6 = wb.create_sheet("Product Performance")
prod_headers = ["Rank","Product Name","Category","Units Sold","Revenue (₹)","Profit (₹)"]
for j, h in enumerate(prod_headers, 1):
    hdr(ws6.cell(1, j), h)

prod_sorted = products.head(15).reset_index(drop=True)
for i, row in enumerate(prod_sorted.itertuples(index=False), 2):
    bg = "F0F0FA" if i % 2 == 0 else "FFFFFF"
    data = [i-1, row.product_name, row.category, row.units, row.revenue, row.profit]
    fmts = [None,None,None,"#,##0","#,##0","#,##0"]
    for j, (v, fmt) in enumerate(zip(data, fmts), 1):
        c = ws6.cell(i, j)
        c.value = v
        c.font = Font(size=10, name="Calibri")
        c.fill = PatternFill("solid", start_color=bg)
        c.alignment = Alignment(horizontal="center")
        c.border = thin_border()
        if fmt: c.number_format = fmt

set_col_widths(ws6, {"A":7,"B":24,"C":14,"D":11,"E":14,"F":14})
print("  ✅ Sheet 6: Product Performance")

# ════════════════════════════════════════════════════════════
# SHEET 7: RFM Customer Table
# ════════════════════════════════════════════════════════════
ws7 = wb.create_sheet("Customer RFM")
rfm = (delivered.groupby(["customer_name","segment","region_name"])
       .agg(frequency=("order_id","nunique"), monetary=("revenue","sum"))
       .reset_index()
       .sort_values("monetary", ascending=False))
rfm["rfm_tier"] = pd.cut(rfm["monetary"],
                          bins=3, labels=["Bronze","Silver","Gold"])

rfm_headers = ["Customer","Segment","Region","Frequency","Monetary (₹)","RFM Tier"]
for j, h in enumerate(rfm_headers, 1):
    hdr(ws7.cell(1, j), h)

tier_colors = {"Gold":"FFD700","Silver":"C0C0C0","Bronze":"CD7F32"}
for i, row in enumerate(rfm.itertuples(index=False), 2):
    bg = tier_colors.get(str(row.rfm_tier), "FFFFFF")
    data = [row.customer_name, row.segment, row.region_name,
            row.frequency, row.monetary, str(row.rfm_tier)]
    fmts = [None,None,None,"#,##0","#,##0",None]
    for j, (v, fmt) in enumerate(zip(data, fmts), 1):
        c = ws7.cell(i, j)
        c.value = v
        c.font = Font(size=10, name="Calibri")
        c.fill = PatternFill("solid", start_color=bg)
        c.alignment = Alignment(horizontal="center")
        c.border = thin_border()
        if fmt: c.number_format = fmt

set_col_widths(ws7, {"A":24,"B":16,"C":12,"D":11,"E":14,"F":11})
print("  ✅ Sheet 7: Customer RFM")

# ── Save workbook ────────────────────────────────────────────
out = EXCEL_OUT / "Business_Analytics_Dashboard.xlsx"
wb.save(out)
print(f"\n🎉 Excel workbook saved → {out}")
